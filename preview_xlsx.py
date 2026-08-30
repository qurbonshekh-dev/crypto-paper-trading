#!/usr/bin/env python3
"""Предпросмотр tracker.xlsx в HTML: стили + посчитанные значения формул."""
import datetime as dt, json, urllib.request, html
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("tracker.xlsx")

# реальные цены, чтобы показать формулы посчитанными
with urllib.request.urlopen("https://api.binance.com/api/v3/ticker/price", timeout=30) as r:
    PX = {t["symbol"][:-4]: float(t["price"]) for t in json.load(r) if t["symbol"].endswith("USDT")}

def money(v): return f"{v:,.2f} $".replace(",", " ")
def pct(v):   return f"{v:+.1f}%"

def compute(ws):
    """Возвращает {(row,col): текст} с посчитанными значениями для листа Сделки."""
    out = {}
    for r in range(5, 105):
        a = ws.cell(r,1).value
        if not a: continue
        coin = ws.cell(r,2).value; bp = ws.cell(r,3).value; amt = ws.cell(r,4).value
        sd = ws.cell(r,5).value;   sp = ws.cell(r,6).value
        qty = amt/bp
        closed = sp is not None
        price = sp if closed else PX.get(coin)
        val = qty*price if price else None
        out[(r,7)] = f"{qty:,.4f}"
        out[(r,8)] = "Закрыт" if closed else "Открыт"
        out[(r,9)] = f"{price:,.2f}" if price and price>=1 else (f"{price:.6f}" if price else "НЕТ ЦЕНЫ")
        out[(r,10)] = money(val) if val else ""
        out[(r,11)] = money(val-amt) if val else ""
        out[(r,12)] = pct((val/amt-1)*100) if val else ""
        d0 = a.date() if hasattr(a,"date") else a
        d1 = (sd.date() if hasattr(sd,"date") else sd) if sd else dt.date.today()
        out[(r,13)] = str((d1-d0).days)
    return out

def render(ws, calc=None, maxr=None, maxc=None):
    maxr = maxr or ws.max_row; maxc = maxc or ws.max_column
    merged = {}
    for m in ws.merged_cells.ranges:
        merged[(m.min_row, m.min_col)] = (m.max_row-m.min_row+1, m.max_col-m.min_col+1)
    skip = set()
    for m in ws.merged_cells.ranges:
        for rr in range(m.min_row, m.max_row+1):
            for cc in range(m.min_col, m.max_col+1):
                if (rr,cc) != (m.min_row, m.min_col): skip.add((rr,cc))
    h = ['<table cellspacing="0">']
    h.append("<colgroup>" + "".join(
        f'<col style="width:{(ws.column_dimensions[get_column_letter(c)].width or 9)*7.5:.0f}px">'
        for c in range(1, maxc+1)) + "</colgroup>")
    for r in range(1, maxr+1):
        if ws.row_dimensions[r].hidden: continue
        ht = ws.row_dimensions[r].height or 19
        h.append(f'<tr style="height:{ht}px">')
        for c in range(1, maxc+1):
            if (r,c) in skip: continue
            cell = ws.cell(r,c)
            v = cell.value
            if calc and (r,c) in calc: v = calc[(r,c)]
            elif isinstance(v,str) and v.startswith("="): v = ""
            elif isinstance(v,(dt.datetime,dt.date)): v = v.strftime("%d.%m.%Y")
            elif isinstance(v,float):
                v = f"{v:,.2f}" if abs(v)>=1 else f"{v:.6f}"
            v = "" if v is None else html.escape(str(v))
            st = []
            def hexof(col):
                try:
                    v = col.rgb
                    v = v if isinstance(v, str) else str(v)
                    return v[-6:] if len(v) >= 6 else None
                except Exception:
                    return None
            f = cell.fill
            bg = hexof(f.fgColor) if f and f.patternType and f.fgColor else None
            if bg and bg != "000000": st.append(f"background:#{bg}")
            fo = cell.font
            if fo:
                fc = hexof(fo.color) if fo.color else None
                if fc and fc != "000000": st.append(f"color:#{fc}")
                if fo.bold: st.append("font-weight:700")
                if fo.italic: st.append("font-style:italic")
                if fo.size: st.append(f"font-size:{fo.size}px")
                if fo.name and "Mono" in str(fo.name): st.append("font-family:'Roboto Mono',monospace")
            al = cell.alignment
            st.append(f"text-align:{al.horizontal or 'left'}")
            st.append("vertical-align:middle")
            if cell.border and cell.border.left and cell.border.left.style:
                st.append("border:1px solid #E1E7EC")
            sp = merged.get((r,c))
            attrs = f' rowspan="{sp[0]}" colspan="{sp[1]}"' if sp else ""
            h.append(f'<td{attrs} style="{";".join(st)}">{v}</td>')
        h.append("</tr>")
    h.append("</table>")
    return "\n".join(h)

sh = wb["Сделки"]
calc = compute(sh)
parts = [f"<h2>{n}</h2>" + render(wb[n], calc if n=="Сделки" else None,
                                  maxr={"Сделки":20,"Цены":16,"Итоги":43}[n],
                                  maxc={"Сделки":14,"Цены":3,"Итоги":4}[n])
         for n in ("Сделки","Цены","Итоги")]
open("preview.html","w").write(f"""<!doctype html><meta charset="utf-8">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&family=Roboto+Mono:wght@400;700&display=swap">
<style>
body{{background:#fff;font-family:Roboto,sans-serif;padding:24px;margin:0}}
h2{{font:700 13px Roboto;color:#6B7A88;text-transform:uppercase;letter-spacing:.1em;margin:28px 0 8px}}
table{{border-collapse:collapse;table-layout:fixed}}
td{{padding:2px 6px;overflow:hidden;white-space:nowrap;font-size:10px}}
</style>{''.join(parts)}""")
print("preview.html готов")
