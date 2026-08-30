#!/usr/bin/env python3
"""Трекер сделок: поиск цены ПО СИМВОЛУ, все расчёты формулами, дашборд."""
import datetime as dt, json, urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ---------- палитра ----------
INK    = "16202B"   # основной текст
SLATE  = "1F2933"   # шапки таблиц
BRONZE = "A9762B"   # акцент: заголовки разделов
CREAM  = "FFF8E7"   # поля для ручного ввода
WHITE  = "FFFFFF"
BAND   = "F4F6F8"   # чередование строк
LINE   = "E1E7EC"
MUTED  = "6B7A88"
UP     = "0F7B4F"
DOWN   = "B03A2E"
WARN   = "9A6B15"
ALARM  = "FDE2DE"

TXT = "Roboto"
NUM = "Roboto Mono"

hair = Side(style="thin", color=LINE)
BOX  = Border(left=hair, right=hair, top=hair, bottom=hair)

TRADES = [
 ("2025-11-07","ETH",3308,50,"2025-11-09",3531,""),
 ("2025-11-07","BNB",947,55,"",None,""),
 ("2025-11-12","SOL",154.88,53,"",None,""),
 ("2025-11-14","SOL",143.76,11,"",None,""),
 ("2025-12-02","BNB",830.29,60,"",895,"дата продажи не была записана"),
 ("2025-12-02","SOL",127,30,"",140,"дата продажи не была записана"),
 ("2025-12-02","ETH",2802,40,"",3194,"дата продажи не была записана"),
 ("2025-12-02","DOGE",0.13576,30,"",0.15,"дата продажи не была записана"),
 ("2026-01-08","XRP",2.162,46,"",None,"была сломанная формула: показывало цену BNB"),
 ("2026-01-10","ZEC",381,50,"",409,"дата продажи не была записана"),
 ("2026-01-16","ZEC",416,40,"",None,"была сломанная формула: показывало цену PYUSD"),
 ("2026-01-16","DOGE",0.14072,40,"",None,""),
 ("2026-01-19","ZEC",365,52,"",None,""),
 ("2026-01-19","ZEC",303,105,"",None,""),
]
ROWS, NCOINS = 100, 90


def fill(c): return PatternFill("solid", fgColor=c)


def coins():
    """Топ по капитализации из CoinGecko: (Название, СИМВОЛ). Тем же источником,
    что и скрипт обновления, — значит символы гарантированно совпадут."""
    req = urllib.request.Request(
        "https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd"
        "&order=market_cap_desc&per_page=250&page=1",
        headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as r:
        data = json.load(r)
    out, seen = [], set()
    for x in data:
        sym = x["symbol"].upper()
        if sym in seen or not x.get("current_price"):
            continue
        seen.add(sym); out.append((x["name"], sym))
        if len(out) >= NCOINS: break
    for _, c, *_ in TRADES:
        if c not in seen: seen.add(c); out.append((c, c))
    return out


def title_bar(ws, row, span, text, sub=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=TXT, bold=True, size=15, color=WHITE)
    c.fill = fill(SLATE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 34
    for k in range(1, span+1):
        ws.cell(row=row, column=k).fill = fill(SLATE)
    if sub:
        s = ws.cell(row=row+1, column=1, value=sub)
        s.font = Font(name=TXT, size=9.5, color=MUTED, italic=True)
        ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=span)
        ws.row_dimensions[row+1].height = 18


def section(ws, row, span, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=TXT, bold=True, size=9.5, color=BRONZE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26


wb = Workbook()

# ═══════════════ ЛИСТ 1: СДЕЛКИ ═══════════════
ws = wb.active
ws.title = "Сделки"
ws.sheet_view.showGridLines = False

title_bar(ws, 1, 14, "  ЖУРНАЛ СДЕЛОК",
          "   Кремовые столбцы заполняешь ты. Белые считаются сами — не трогай их. "
          "Цена ищется по символу монеты, поэтому сортировка списка цен ничего не ломает.")

HDR = ["Дата\nпокупки","Монета","Цена\nпокупки","Сумма\nпокупки","Дата\nпродажи",
       "Цена\nпродажи","Кол-во\nмонет","Статус","Цена\nсейчас","Стоимость\nсейчас",
       "Разница $","Разница %","Дней","Комментарий"]
IN_COLS = {1,2,3,4,5,6,14}
HR = 4
for i, h in enumerate(HDR, 1):
    c = ws.cell(row=HR, column=i, value=h)
    c.font = Font(name=TXT, bold=True, size=9.5, color=WHITE)
    c.fill = fill(SLATE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BOX
ws.row_dimensions[HR].height = 40
ws.freeze_panes = f"A{HR+1}"

first, last = HR+1, HR+ROWS
for r in range(first, last+1):
    d = TRADES[r-first] if r-first < len(TRADES) else None
    if d:
        ws.cell(row=r, column=1, value=dt.date.fromisoformat(d[0]))
        ws.cell(row=r, column=2, value=d[1])
        ws.cell(row=r, column=3, value=d[2])
        ws.cell(row=r, column=4, value=d[3])
        if d[4]: ws.cell(row=r, column=5, value=dt.date.fromisoformat(d[4]))
        if d[5] is not None: ws.cell(row=r, column=6, value=d[5])
        if d[6]: ws.cell(row=r, column=14, value=d[6])
    ws.cell(row=r, column=7,  value=f'=IF(OR($C{r}="",$D{r}=""),"",$D{r}/$C{r})')
    ws.cell(row=r, column=8,  value=f'=IF($A{r}="","",IF($F{r}<>"","Закрыт","Открыт"))')
    ws.cell(row=r, column=9,  value=(
        f'=IF($B{r}="","",IF($F{r}<>"",$F{r},'
        f'IFERROR(INDEX(Цены!$C:$C,MATCH($B{r},Цены!$B:$B,0)),"НЕТ ЦЕНЫ")))'))
    ws.cell(row=r, column=10, value=f'=IF(OR($G{r}="",NOT(ISNUMBER($I{r}))),"",$G{r}*$I{r})')
    ws.cell(row=r, column=11, value=f'=IF($J{r}="","",$J{r}-$D{r})')
    ws.cell(row=r, column=12, value=f'=IF(OR($K{r}="",$D{r}=""),"",$K{r}/$D{r})')
    ws.cell(row=r, column=13, value=f'=IF($A{r}="","",IF($E{r}<>"",$E{r}-$A{r},TODAY()-$A{r}))')
    ws.row_dimensions[r].height = 21
    for c in range(1, 15):
        cell = ws.cell(row=r, column=c)
        cell.border = BOX
        cell.fill = fill(CREAM if c in IN_COLS else WHITE)
        mono = c in (3,4,6,7,9,10,11,12,13)
        cell.font = Font(name=NUM if mono else TXT, size=10, color=INK)
        if c in (1,5):
            cell.number_format = "DD.MM.YYYY"; cell.alignment = Alignment(horizontal="center")
        elif c == 2:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name=TXT, size=10, bold=True, color=INK)
        elif c == 8:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name=TXT, size=9, bold=True)
        elif c in (3,6,9):
            cell.number_format = "#,##0.00######"; cell.alignment = Alignment(horizontal="right")
        elif c == 7:
            cell.number_format = "#,##0.0000####"; cell.alignment = Alignment(horizontal="right")
        elif c in (4,10,11):
            cell.number_format = '#,##0.00" $"'; cell.alignment = Alignment(horizontal="right")
        elif c == 12:
            cell.number_format = "+0.0%;-0.0%"; cell.alignment = Alignment(horizontal="right")
        elif c == 13:
            cell.number_format = "0"; cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left", indent=1)
            cell.font = Font(name=TXT, size=9, color=MUTED, italic=True)

rng = f"A{first}:N{last}"
# чередование строк — только там, где строка заполнена
ws.conditional_formatting.add(rng, FormulaRule(
    formula=[f'AND(ISEVEN(ROW()),$A{first}<>"")'], fill=fill(BAND)))
for col in ("K","L"):
    cr = f"{col}{first}:{col}{last}"
    ws.conditional_formatting.add(cr, CellIsRule(operator="greaterThan", formula=["0"],
        font=Font(name=NUM, size=10, color=UP, bold=True)))
    ws.conditional_formatting.add(cr, CellIsRule(operator="lessThan", formula=["0"],
        font=Font(name=NUM, size=10, color=DOWN, bold=True)))
ws.conditional_formatting.add(f"H{first}:H{last}", FormulaRule(
    formula=[f'$H{first}="Открыт"'], fill=fill("FFF2D6"),
    font=Font(name=TXT, size=9, bold=True, color=WARN)))
ws.conditional_formatting.add(f"H{first}:H{last}", FormulaRule(
    formula=[f'$H{first}="Закрыт"'], fill=fill("E8F0EB"),
    font=Font(name=TXT, size=9, bold=True, color=UP)))
# отсутствующий символ виден сразу, а не подменяется чужой ценой
ws.conditional_formatting.add(f"I{first}:N{last}", FormulaRule(
    formula=[f'$I{first}="НЕТ ЦЕНЫ"'], fill=fill(ALARM),
    font=Font(name=NUM, size=10, color=DOWN, bold=True)))

dv = DataValidation(type="list", formula1=f"Цены!$B$5:$B${4+NCOINS+6}",
                    allow_blank=True, showDropDown=False,
                    error="Такой монеты нет в листе «Цены». Добавь её туда, "
                          "иначе цена не подтянется.",
                    errorTitle="Неизвестная монета", promptTitle="Монета",
                    prompt="Выбери из списка — так цена точно подтянется")
ws.add_data_validation(dv)
dv.add(f"B{first}:B{last}")

for c, w in zip("ABCDEFGHIJKLMN", (12,10,13,13,12,13,13,10,13,14,12,11,7,40)):
    ws.column_dimensions[c].width = w
ws.auto_filter.ref = f"A{HR}:N{last}"

# ═══════════════ ЛИСТ 2: ЦЕНЫ ═══════════════
p = wb.create_sheet("Цены")
p.sheet_view.showGridLines = False
title_bar(p, 1, 4, "  ЦЕНЫ")
p["A2"] = "   Обновляет скрипт раз в час. Вручную ничего не меняй."
p["A2"].font = Font(name=TXT, size=9.5, color=MUTED, italic=True)
p.merge_cells("A2:D2")

p["A3"] = "Обновлено:"
p["A3"].font = Font(name=TXT, bold=True, size=10, color=INK)
p["B3"] = "—"
p["B3"].font = Font(name=NUM, size=10)
p["B3"].number_format = "DD.MM.YYYY HH:MM"
p["B3"].alignment = Alignment(horizontal="center")
p["C3"] = ('=IF(B3="—","скрипт ещё не запускался",'
           'IF(NOW()-B3>0.5,"ЦЕНЫ УСТАРЕЛИ","цены свежие"))')
p["C3"].font = Font(name=TXT, bold=True, size=10)
p["C3"].alignment = Alignment(horizontal="center")
p.conditional_formatting.add("C3", FormulaRule(
    formula=['ISNUMBER(SEARCH("УСТАРЕЛИ",$C$3))'], fill=fill(ALARM),
    font=Font(name=TXT, bold=True, size=10, color=DOWN)))
p.conditional_formatting.add("C3", FormulaRule(
    formula=['$C$3="цены свежие"'],
    font=Font(name=TXT, bold=True, size=10, color=UP)))
p["D3"] = "источник: —"
p["D3"].font = Font(name=TXT, size=9, color=MUTED, italic=True)
p["D3"].alignment = Alignment(horizontal="left", indent=1)
p.row_dimensions[3].height = 24

for i, h in enumerate(["Название","Символ","Цена USD"], 1):
    c = p.cell(row=4, column=i, value=h)
    c.font = Font(name=TXT, bold=True, size=9.5, color=WHITE)
    c.fill = fill(SLATE); c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center")
p.row_dimensions[4].height = 24
p.freeze_panes = "A5"

lst = coins()
for i, pair in enumerate(lst + [("", "")]*6, 5):
    name, cn = pair
    p.cell(row=i, column=1, value=name)
    p.cell(row=i, column=2, value=cn)
    if cn: p.cell(row=i, column=3, value=0)
    for col in (1, 2, 3):
        cell = p.cell(row=i, column=col)
        cell.border = BOX; cell.fill = fill(WHITE)
        cell.font = Font(name=NUM if col == 3 else TXT, size=10, color=INK)
        cell.alignment = Alignment(horizontal="right" if col == 3 else
                                   ("center" if col == 2 else "left"), indent=0 if col==3 else 1)
    p.cell(row=i, column=3).number_format = "#,##0.00########"
    p.row_dimensions[i].height = 19
p.conditional_formatting.add(f"A5:C{4+len(lst)+6}", FormulaRule(
    formula=["AND(ISEVEN(ROW()),$B5<>\"\")"], fill=fill(BAND)))
for c, w in zip("ABCD", (26, 14, 20, 26)): p.column_dimensions[c].width = w

# ═══════════════ ЛИСТ 3: ИТОГИ ═══════════════
t = wb.create_sheet("Итоги")
t.sheet_view.showGridLines = False
S = "Сделки"
title_bar(t, 1, 4, "  ИТОГИ ПОРТФЕЛЯ",
          "   Считается само из листа «Сделки». Открытые позиции — по текущей цене.")

def kv(row, label, formula, fmt, big=False, color=False, tint=None):
    a = t.cell(row=row, column=1, value=label)
    a.font = Font(name=TXT, size=11 if big else 10, color=INK if big else MUTED,
                  bold=big)
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    b = t.cell(row=row, column=2, value=formula)
    b.font = Font(name=NUM, size=16 if big else 11, bold=True, color=INK)
    b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    b.number_format = {"money": '#,##0.00" $"', "pct": "+0.0%;-0.0%", "num": "0"}[fmt]
    for c in (1, 2):
        t.cell(row=row, column=c).border = BOX
        t.cell(row=row, column=c).fill = fill(tint or WHITE)
    t.row_dimensions[row].height = 30 if big else 22
    if color:
        for rule, col in ((("greaterThan", ["0"]), UP), (("lessThan", ["0"]), DOWN)):
            t.conditional_formatting.add(f"B{row}", CellIsRule(
                operator=rule[0], formula=rule[1],
                font=Font(name=NUM, size=16 if big else 11, bold=True, color=col)))

section(t, 4, 4, "ВСЕГО")
HERO = "FBFAF6"
kv(5, "Вложено",           f'=B11+B18', "money", big=True, tint=HERO)
kv(6, "Стоимость сейчас",  f'=B12+B19', "money", big=True, tint=HERO)
kv(7, "Прибыль",           f'=B6-B5',   "money", big=True, color=True, tint=HERO)
kv(8, "в процентах",       f'=IF(B5=0,"",B7/B5)', "pct", big=True, color=True, tint=HERO)

section(t, 10, 4, "ОТКРЫТЫЕ ПОЗИЦИИ")
kv(11, "Вложено",              f'=SUMIF({S}!$H:$H,"Открыт",{S}!$D:$D)', "money")
kv(12, "Стоимость сейчас",     f'=SUMIF({S}!$H:$H,"Открыт",{S}!$J:$J)', "money")
kv(13, "Нереализованный P&L",  f'=B12-B11', "money", color=True)
kv(14, "в процентах",          f'=IF(B11=0,"",B13/B11)', "pct", color=True)
kv(15, "Позиций открыто",      f'=COUNTIF({S}!$H:$H,"Открыт")', "num")

section(t, 17, 4, "ЗАКРЫТЫЕ СДЕЛКИ")
kv(18, "Вложено",             f'=SUMIF({S}!$H:$H,"Закрыт",{S}!$D:$D)', "money")
kv(19, "Получено",            f'=SUMIF({S}!$H:$H,"Закрыт",{S}!$J:$J)', "money")
kv(20, "Реализованный P&L",   f'=B19-B18', "money", color=True)
kv(21, "в процентах",         f'=IF(B18=0,"",B20/B18)', "pct", color=True)
kv(22, "Сделок закрыто",      f'=COUNTIF({S}!$H:$H,"Закрыт")', "num")
kv(23, "Из них прибыльных",   f'=COUNTIFS({S}!$H:$H,"Закрыт",{S}!$K:$K,">0")', "num")
kv(24, "Winrate",             f'=IF(B22=0,"",B23/B22)', "pct")
t["B24"].number_format = "0.0%"

section(t, 26, 4, "ПО МОНЕТАМ · открытые позиции")
for i, h in enumerate(["Монета","Вложено","Сейчас","P&L"], 1):
    c = t.cell(row=27, column=i, value=h)
    c.font = Font(name=TXT, bold=True, size=9.5, color=WHITE)
    c.fill = fill(SLATE); c.border = BOX
    c.alignment = Alignment(horizontal="center", vertical="center")
t.row_dimensions[27].height = 24

uniq = []
for _, c, *_ in TRADES:
    if c not in uniq: uniq.append(c)
for i in range(14):
    r = 28 + i
    if i < len(uniq):
        cc = t.cell(row=r, column=1, value=uniq[i])
        cc.fill = fill(CREAM)
    else:
        t.cell(row=r, column=1).fill = fill(CREAM)
    t.cell(row=r, column=1).font = Font(name=TXT, size=10, bold=True, color=INK)
    t.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    t.cell(row=r, column=2, value=f'=IF($A{r}="","",SUMIFS({S}!$D:$D,{S}!$B:$B,$A{r},{S}!$H:$H,"Открыт"))')
    t.cell(row=r, column=3, value=f'=IF($A{r}="","",SUMIFS({S}!$J:$J,{S}!$B:$B,$A{r},{S}!$H:$H,"Открыт"))')
    t.cell(row=r, column=4, value=f'=IF(OR($A{r}="",$B{r}=0),"",$C{r}-$B{r})')
    for col in range(2, 5):
        cell = t.cell(row=r, column=col)
        cell.number_format = '#,##0.00" $"'
        cell.font = Font(name=NUM, size=10, color=INK)
        cell.alignment = Alignment(horizontal="right", indent=1)
    for col in range(1, 5):
        t.cell(row=r, column=col).border = BOX
        if col > 1: t.cell(row=r, column=col).fill = fill(WHITE)
    t.row_dimensions[r].height = 20
    for op, col in (("greaterThan", UP), ("lessThan", DOWN)):
        t.conditional_formatting.add(f"D{r}", CellIsRule(operator=op, formula=["0"],
            font=Font(name=NUM, size=10, bold=True, color=col)))
t.cell(row=28+14, column=1, value="↑ впиши символ, строка посчитается сама").font = \
    Font(name=TXT, size=9, color=MUTED, italic=True)
t.merge_cells(start_row=28+14, start_column=1, end_row=28+14, end_column=4)

for c, w in zip("ABCD", (30, 20, 16, 16)): t.column_dimensions[c].width = w

wb.save("tracker.xlsx")
print("tracker.xlsx собран")
